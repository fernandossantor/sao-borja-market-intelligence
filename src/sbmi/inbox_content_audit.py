"""Auditoria de conteúdo local para as tabelas capturadas de ``raw/new_files``."""

from __future__ import annotations

import csv
import hashlib
import math
import re
import unicodedata
from collections import Counter
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import date, datetime
from itertools import combinations
from pathlib import Path, PurePosixPath

import pandas as pd
from openpyxl import load_workbook

from sbmi.inbox_profile import normalize_label
from sbmi.inbox_structure_triage import source_from_path

DATE_HEADER_PRIORITY = (
    "mes_ano",
    "data",
    "competencia",
    "ano_mes",
    "mes_competencia",
    "data_pagamento",
    "data_lancamento",
)
PORTUGUESE_MONTHS = {
    "jan": 1,
    "janeiro": 1,
    "fev": 2,
    "fevereiro": 2,
    "mar": 3,
    "marco": 3,
    "abr": 4,
    "abril": 4,
    "mai": 5,
    "maio": 5,
    "jun": 6,
    "junho": 6,
    "jul": 7,
    "julho": 7,
    "ago": 8,
    "agosto": 8,
    "set": 9,
    "setembro": 9,
    "out": 10,
    "outubro": 10,
    "nov": 11,
    "novembro": 11,
    "dez": 12,
    "dezembro": 12,
}
MONTH_YEAR_PATTERN = re.compile(r"^(?P<month>\d{1,2})[/-](?P<year>\d{4})$")
YEAR_MONTH_PATTERN = re.compile(r"^(?P<year>\d{4})[/-](?P<month>\d{1,2})$")
TEXT_MONTH_PATTERN = re.compile(
    r"^(?P<month>[A-Za-zÀ-ÿ]+)[\s./-]+(?P<year>\d{4})$",
    flags=re.IGNORECASE,
)


@dataclass(frozen=True)
class LoadedTable:
    relative_path: str
    sheet_name: str
    sheet_index: int
    source_declared: str
    headers: tuple[str, ...]
    normalized_headers: tuple[str, ...]
    rows: tuple[tuple[object, ...], ...]


@dataclass(frozen=True)
class ContentAuditResult:
    table_summary: pd.DataFrame
    federal_overlap_candidates: pd.DataFrame
    audit_summary: pd.DataFrame


def _ascii_fold(value: str) -> str:
    decomposed = unicodedata.normalize("NFKD", value)
    return "".join(
        char for char in decomposed if not unicodedata.combining(char)
    )


def canonical_value(value: object) -> str:
    """Converte um valor observado em representação estável."""
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return format(value, ".15g")
    return re.sub(r"\s+", " ", str(value).strip())


def normalized_value(value: object) -> str:
    """Normaliza um valor para localizar candidatos de igualdade."""
    canonical = canonical_value(value)
    return re.sub(r"\s+", " ", _ascii_fold(canonical).casefold()).strip()


def _row_hash(row: Iterable[object], *, normalized: bool) -> str:
    converter = normalized_value if normalized else canonical_value
    payload = "\x1f".join(converter(value) for value in row)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _multiset_hash(row_hashes: Iterable[str]) -> str:
    payload = "\n".join(sorted(row_hashes))
    return hashlib.sha256(payload.encode("ascii")).hexdigest()


def _sequence_hash(row_hashes: Iterable[str]) -> str:
    payload = "\n".join(row_hashes)
    return hashlib.sha256(payload.encode("ascii")).hexdigest()


def parse_temporal_value(value: object) -> str | None:
    """Interpreta datas somente em coluna temporal identificada."""
    if value is None:
        return None
    if isinstance(value, pd.Timestamp):
        return value.date().isoformat()
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = canonical_value(value)
    if not text:
        return None

    month_year = MONTH_YEAR_PATTERN.fullmatch(text)
    if month_year:
        month = int(month_year.group("month"))
        year = int(month_year.group("year"))
        if 1 <= month <= 12:
            return date(year, month, 1).isoformat()

    year_month = YEAR_MONTH_PATTERN.fullmatch(text)
    if year_month:
        month = int(year_month.group("month"))
        year = int(year_month.group("year"))
        if 1 <= month <= 12:
            return date(year, month, 1).isoformat()

    text_month = TEXT_MONTH_PATTERN.fullmatch(text)
    if text_month:
        month_token = normalize_label(text_month.group("month"))
        month = PORTUGUESE_MONTHS.get(month_token)
        if month is not None:
            year = int(text_month.group("year"))
            return date(year, month, 1).isoformat()

    parsed = pd.to_datetime(text, errors="coerce", dayfirst=True)
    if pd.isna(parsed):
        return None
    return parsed.date().isoformat()


def _decode_delimited(path: Path) -> tuple[str, str]:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return raw.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError(
        "utf-8",
        raw,
        0,
        1,
        "Não foi possível decodificar o arquivo",
    )


def _load_delimited(
    path: Path,
    header_row: int,
    width: int,
) -> tuple[list[str], list[list[object]]]:
    text, _ = _decode_delimited(path)
    sample = text[:65536]
    default = "\t" if path.suffix.lower() == ".tsv" else ","
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        delimiter = default
    parsed = list(csv.reader(text.splitlines(), delimiter=delimiter))
    header = parsed[header_row - 1][:width]
    rows = [row[:width] for row in parsed[header_row:]]
    return header, rows


def _load_excel(
    path: Path,
    sheet_name: str,
    header_row: int,
    width: int,
) -> tuple[list[str], list[list[object]]]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        worksheet = workbook[sheet_name]
        header: list[str] = []
        rows: list[list[object]] = []
        iterator = worksheet.iter_rows(max_col=width)
        for row_index, row in enumerate(iterator, start=1):
            values = [cell.value for cell in row]
            if row_index == header_row:
                header = [canonical_value(value) for value in values]
            elif row_index > header_row:
                rows.append(values)
        return header, rows
    finally:
        workbook.close()


def _clean_rows(
    rows: Iterable[Iterable[object]],
    width: int,
) -> tuple[tuple[object, ...], ...]:
    cleaned: list[tuple[object, ...]] = []
    for row in rows:
        row_values = list(row)[:width]
        row_values.extend([None] * max(0, width - len(row_values)))
        values = tuple(row_values)
        if any(canonical_value(value) for value in values):
            cleaned.append(values)
    return tuple(cleaned)


def load_profiled_tables(
    snapshot_path: Path,
    sheet_profile: pd.DataFrame,
) -> tuple[list[LoadedTable], list[dict[str, object]]]:
    """Carrega tabelas usando as linhas de cabeçalho do perfil."""
    required = {
        "relative_path",
        "sheet_name",
        "sheet_index",
        "header_row_candidate_estimate",
        "observed_max_column",
    }
    missing = required.difference(sheet_profile.columns)
    if missing:
        raise ValueError(f"Colunas obrigatórias ausentes: {sorted(missing)}")

    root = snapshot_path.expanduser().resolve()
    loaded: list[LoadedTable] = []
    errors: list[dict[str, object]] = []

    for row in sheet_profile.itertuples(index=False):
        relative_path = str(row.relative_path)
        path = root / PurePosixPath(relative_path)
        try:
            header_row = int(row.header_row_candidate_estimate)
            width = int(row.observed_max_column)
            if header_row < 1 or width < 1:
                raise ValueError("Cabeçalho ou largura inválidos no perfil.")
            if path.suffix.lower() in {".xlsx", ".xlsm"}:
                headers, rows = _load_excel(
                    path,
                    str(row.sheet_name),
                    header_row,
                    width,
                )
            elif path.suffix.lower() in {".csv", ".tsv", ".txt"}:
                headers, rows = _load_delimited(path, header_row, width)
            else:
                raise ValueError(f"Formato não suportado: {path.suffix}")

            normalized_headers = tuple(
                normalize_label(value) or f"column_{index}"
                for index, value in enumerate(headers, start=1)
            )
            loaded.append(
                LoadedTable(
                    relative_path=relative_path,
                    sheet_name=str(row.sheet_name),
                    sheet_index=int(row.sheet_index),
                    source_declared=source_from_path(relative_path),
                    headers=tuple(headers),
                    normalized_headers=normalized_headers,
                    rows=_clean_rows(rows, width),
                )
            )
        except Exception as exc:  # noqa: BLE001
            errors.append(
                {
                    "relative_path": relative_path,
                    "sheet_name": str(row.sheet_name),
                    "sheet_index": int(row.sheet_index),
                    "error_type": type(exc).__name__,
                    "error_message": str(exc)[:500],
                }
            )
    return loaded, errors


def _date_header(table: LoadedTable) -> tuple[int | None, str]:
    for candidate in DATE_HEADER_PRIORITY:
        if candidate in table.normalized_headers:
            return table.normalized_headers.index(candidate), candidate
    return None, ""


def build_table_summary(tables: Iterable[LoadedTable]) -> pd.DataFrame:
    """Resume linhas, duplicidades internas e períodos por tabela."""
    records: list[dict[str, object]] = []
    for table in tables:
        strict_hashes = [
            _row_hash(row, normalized=False) for row in table.rows
        ]
        normalized_hashes = [
            _row_hash(row, normalized=True) for row in table.rows
        ]
        normalized_counts = Counter(normalized_hashes)
        date_index, date_header = _date_header(table)
        date_values = (
            [row[date_index] for row in table.rows]
            if date_index is not None
            else []
        )
        nonblank_dates = [
            value for value in date_values if canonical_value(value)
        ]
        parsed_dates = [
            parsed
            for value in nonblank_dates
            if (parsed := parse_temporal_value(value))
        ]

        records.append(
            {
                "relative_path": table.relative_path,
                "source_declared": table.source_declared,
                "sheet_name": table.sheet_name,
                "sheet_index": table.sheet_index,
                "rows_observed": len(table.rows),
                "unique_rows_normalized": len(normalized_counts),
                "duplicate_rows_within_file": (
                    len(table.rows) - len(normalized_counts)
                ),
                "strict_sequence_sha256": _sequence_hash(strict_hashes),
                "normalized_sequence_sha256": _sequence_hash(
                    normalized_hashes
                ),
                "normalized_multiset_sha256": _multiset_hash(
                    normalized_hashes
                ),
                "date_header_observed": date_header,
                "date_values_nonblank": len(nonblank_dates),
                "date_values_parsed": len(parsed_dates),
                "date_parse_failures": len(nonblank_dates) - len(parsed_dates),
                "date_parse_rate": (
                    round(len(parsed_dates) / len(nonblank_dates), 6)
                    if nonblank_dates
                    else None
                ),
                "period_min_observed": (
                    min(parsed_dates) if parsed_dates else None
                ),
                "period_max_observed": (
                    max(parsed_dates) if parsed_dates else None
                ),
                "normalized_row_hashes": "|".join(normalized_hashes),
            }
        )
    return pd.DataFrame(records)


def build_federal_overlap_candidates(summary: pd.DataFrame) -> pd.DataFrame:
    """Compara o conteúdo normalizado das tabelas federais."""
    columns = [
        "left_path",
        "right_path",
        "left_rows",
        "right_rows",
        "shared_unique_rows",
        "jaccard_row_similarity",
        "left_containment",
        "right_containment",
        "same_normalized_multiset",
        "candidate_class",
        "left_period_min",
        "left_period_max",
        "right_period_min",
        "right_period_max",
    ]
    if summary.empty:
        return pd.DataFrame(columns=columns)

    federal = summary.loc[summary["source_declared"].eq("Federal")].copy()
    records: list[dict[str, object]] = []

    for left, right in combinations(federal.itertuples(index=False), 2):
        left_hashes = [
            value
            for value in str(left.normalized_row_hashes).split("|")
            if value
        ]
        right_hashes = [
            value
            for value in str(right.normalized_row_hashes).split("|")
            if value
        ]
        left_set = set(left_hashes)
        right_set = set(right_hashes)
        intersection = len(left_set & right_set)
        if intersection == 0:
            continue

        union = len(left_set | right_set)
        left_containment = intersection / len(left_set) if left_set else 0.0
        right_containment = (
            intersection / len(right_set) if right_set else 0.0
        )
        identical_multiset = (
            left.normalized_multiset_sha256
            == right.normalized_multiset_sha256
        )
        if identical_multiset:
            candidate_class = "IDENTICAL_NORMALIZED_CONTENT"
        elif left_containment == 1.0:
            candidate_class = "LEFT_CONTAINED_IN_RIGHT"
        elif right_containment == 1.0:
            candidate_class = "RIGHT_CONTAINED_IN_LEFT"
        else:
            candidate_class = "PARTIAL_ROW_OVERLAP"

        records.append(
            {
                "left_path": left.relative_path,
                "right_path": right.relative_path,
                "left_rows": int(left.rows_observed),
                "right_rows": int(right.rows_observed),
                "shared_unique_rows": intersection,
                "jaccard_row_similarity": (
                    round(intersection / union, 6) if union else 0.0
                ),
                "left_containment": round(left_containment, 6),
                "right_containment": round(right_containment, 6),
                "same_normalized_multiset": bool(identical_multiset),
                "candidate_class": candidate_class,
                "left_period_min": left.period_min_observed,
                "left_period_max": left.period_max_observed,
                "right_period_min": right.period_min_observed,
                "right_period_max": right.period_max_observed,
            }
        )

    if not records:
        return pd.DataFrame(columns=columns)
    return (
        pd.DataFrame(records)
        .sort_values(
            [
                "same_normalized_multiset",
                "jaccard_row_similarity",
                "shared_unique_rows",
            ],
            ascending=[False, False, False],
        )
        .reset_index(drop=True)
    )


def build_audit_summary(
    table_summary: pd.DataFrame,
    overlap_candidates: pd.DataFrame,
    errors: list[dict[str, object]],
) -> pd.DataFrame:
    """Produz indicadores agregados e registra sua natureza."""
    if table_summary.empty:
        federal_tables = 0
        duplicate_tables = 0
        temporal_tables = 0
        date_failures = 0
    else:
        federal_tables = int(
            table_summary["source_declared"].eq("Federal").sum()
        )
        duplicate_tables = int(
            table_summary["duplicate_rows_within_file"].gt(0).sum()
        )
        temporal_tables = int(
            table_summary["date_header_observed"].fillna("").ne("").sum()
        )
        date_failures = int(table_summary["date_parse_failures"].sum())

    classes = (
        overlap_candidates["candidate_class"].value_counts()
        if not overlap_candidates.empty
        else {}
    )
    indicators = [
        ("tables_loaded", len(table_summary), "observed"),
        ("tables_error", len(errors), "observed"),
        ("federal_tables", federal_tables, "observed"),
        (
            "tables_with_internal_duplicate_rows",
            duplicate_tables,
            "calculated",
        ),
        ("tables_with_date_header", temporal_tables, "observed"),
        ("date_parse_failures_total", date_failures, "calculated"),
        ("federal_overlap_pairs", len(overlap_candidates), "calculated"),
        (
            "identical_normalized_content_pairs",
            int(classes.get("IDENTICAL_NORMALIZED_CONTENT", 0)),
            "calculated",
        ),
        (
            "containment_pairs",
            int(classes.get("LEFT_CONTAINED_IN_RIGHT", 0))
            + int(classes.get("RIGHT_CONTAINED_IN_LEFT", 0)),
            "calculated",
        ),
        (
            "partial_row_overlap_pairs",
            int(classes.get("PARTIAL_ROW_OVERLAP", 0)),
            "calculated",
        ),
    ]
    return pd.DataFrame(
        indicators,
        columns=["indicator", "value", "nature"],
    )


def audit_snapshot_content(
    snapshot_path: Path,
    sheet_profile: pd.DataFrame,
) -> tuple[ContentAuditResult, pd.DataFrame]:
    """Executa auditoria de conteúdo sem modificar a captura local."""
    tables, errors = load_profiled_tables(snapshot_path, sheet_profile)
    table_summary = build_table_summary(tables)
    overlap = build_federal_overlap_candidates(table_summary)
    summary = build_audit_summary(table_summary, overlap, errors)
    public_summary = table_summary.drop(
        columns=["normalized_row_hashes"],
        errors="ignore",
    )
    return (
        ContentAuditResult(
            table_summary=public_summary,
            federal_overlap_candidates=overlap,
            audit_summary=summary,
        ),
        pd.DataFrame(errors),
    )
