"""Auditoria de metadados e evidências de proveniência das planilhas censitárias."""

from __future__ import annotations

import json
import re
import shutil
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path, PurePosixPath
from urllib.parse import urlparse

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

from sbmi.base_territorial_coverage import normalize_text
from sbmi.demography_census_comparison import canonical_value, normalize_header

LINEAGE_REQUIRED_COLUMNS = {
    "dataset_identity",
    "raw_source_count",
    "raw_source_paths",
    "lineage_match_status",
}
MANIFEST_REQUIRED_COLUMNS = {
    "relative_path",
    "expected_size_bytes",
    "downloaded_size_bytes",
    "expected_sha256",
    "local_sha256",
    "verification_status",
}
MATCHED_STATUS = "MATCHED_ONE_TO_ONE_BY_NAME"
URL_PATTERN = re.compile(r"https?://[^\s<>\]\[\)\(\"']+", re.IGNORECASE)
SOURCE_TERMS = (
    "fonte",
    "source",
    "origem dos dados",
    "ibge",
    "sidra",
    "instituto brasileiro de geografia e estatistica",
)
PERIOD_TERMS = (
    "censo 2022",
    "ano de referencia",
    "data de referencia",
    "periodo de referencia",
    "2022",
)
UNIT_TERMS = (
    "unidade",
    "percentual",
    "porcentagem",
    "km2",
    "km 2",
    "hab km2",
    "habitantes por km2",
)
GEOGRAPHY_TERMS = (
    "sao borja",
    "4318002",
    "codigo ibge",
    "municipio",
)
INSTITUTION_TERMS = (
    "ibge",
    "sidra",
    "instituto brasileiro de geografia e estatistica",
)
DOCUMENT_PROPERTY_FIELDS = (
    "creator",
    "lastModifiedBy",
    "title",
    "subject",
    "description",
    "keywords",
    "category",
    "identifier",
    "language",
    "version",
    "contentStatus",
)


@dataclass(frozen=True)
class CensusProvenanceAuditResult:
    workbooks: pd.DataFrame
    sheets: pd.DataFrame
    columns: pd.DataFrame
    evidence: pd.DataFrame
    summary: pd.DataFrame


def _safe_snapshot_file(root: Path, relative_path: object) -> Path:
    normalized = PurePosixPath(str(relative_path or "").strip("/"))
    if not normalized.parts or normalized.is_absolute():
        raise ValueError(f"Caminho de captura inválido: {relative_path!r}")
    candidate = root.joinpath(*normalized.parts).resolve()
    resolved_root = root.resolve()
    try:
        candidate.relative_to(resolved_root)
    except ValueError as exc:
        raise ValueError(f"Caminho fora da captura: {relative_path}") from exc
    if not candidate.is_file():
        raise FileNotFoundError(f"Arquivo da captura não encontrado: {candidate}")
    return candidate


def _single_path(value: object, expected_count: object) -> str:
    count = int(expected_count)
    paths = [part.strip() for part in str(value or "").split("|") if part.strip()]
    if count != 1 or len(paths) != 1:
        raise ValueError(
            "Fonte bruta não corresponde a um caminho único: "
            f"count={count}, paths={paths}"
        )
    return paths[0]


def _iso(value: object) -> str:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value or "").strip()


def _text(value: object, limit: int = 500) -> str:
    compact = " ".join(str(value or "").split())
    return compact[:limit]


def _domain(value: str) -> str:
    parsed = urlparse(value)
    return parsed.netloc.casefold().removeprefix("www.")


def _contains_any(normalized: str, terms: tuple[str, ...]) -> bool:
    return any(normalize_text(term) in normalized for term in terms)


def _evidence_kind(text: str) -> list[str]:
    normalized = normalize_text(text)
    kinds: list[str] = []
    if _contains_any(normalized, SOURCE_TERMS):
        kinds.append("SOURCE_LABEL")
    if _contains_any(normalized, PERIOD_TERMS):
        kinds.append("PERIOD_LABEL")
    if _contains_any(normalized, UNIT_TERMS):
        kinds.append("UNIT_LABEL")
    if _contains_any(normalized, GEOGRAPHY_TERMS):
        kinds.append("GEOGRAPHY_LABEL")
    if _contains_any(normalized, INSTITUTION_TERMS):
        kinds.append("INSTITUTION_LABEL")
    return kinds


def _authority_implication(kind: str) -> str:
    if kind in {"CELL_URL", "HYPERLINK_TARGET"}:
        return "DOMAIN_REQUIRES_EXTERNAL_VERIFICATION"
    if kind in {"SOURCE_LABEL", "INSTITUTION_LABEL"}:
        return "SOURCE_CLAIM_REQUIRES_EXTERNAL_VERIFICATION"
    if kind.startswith("DOCUMENT_PROPERTY_"):
        return "DOCUMENT_METADATA_DOES_NOT_ESTABLISH_AUTHORITY"
    return "CONTEXT_HINT_DOES_NOT_ESTABLISH_AUTHORITY"


def _unit_hint(header: object) -> str:
    normalized = normalize_header(header)
    if any(token in normalized for token in ("porcentagem", "percentual", "percent")):
        return "PERCENT"
    if "densidade" in normalized and any(
        token in normalized for token in ("km2", "km_2")
    ):
        return "INHABITANTS_PER_SQUARE_KILOMETER"
    if "area" in normalized and any(token in normalized for token in ("km2", "km_2")):
        return "SQUARE_KILOMETERS"
    if any(
        token in normalized
        for token in (
            "populacao",
            "pessoas",
            "domicilios",
            "habitantes",
            "quantidade",
            "total",
        )
    ):
        return "COUNT_OR_TOTAL_REQUIRES_CONFIRMATION"
    if "taxa" in normalized:
        return "RATE_REQUIRES_DEFINITION"
    return "UNSPECIFIED"


def _first_nonempty_row(sheet: object) -> int | None:
    for row_number in range(1, int(sheet.max_row) + 1):
        if any(sheet.cell(row=row_number, column=column).value is not None for column in range(1, int(sheet.max_column) + 1)):
            return row_number
    return None


def _cell_text(cell: Cell) -> str:
    return _text(cell.value)


def _add_evidence(
    accumulator: dict[tuple[str, str, str, str], dict[str, object]],
    *,
    dataset_identity: str,
    raw_relative_path: str,
    sheet_name: str,
    cell_coordinate: str,
    kind: str,
    evidence_text: str,
    detected_domain: str = "",
) -> None:
    compact = _text(evidence_text)
    if not compact:
        return
    key = (dataset_identity, kind, compact, detected_domain)
    if key not in accumulator:
        accumulator[key] = {
            "dataset_identity": dataset_identity,
            "raw_relative_path": raw_relative_path,
            "sheet_name": sheet_name,
            "cell_coordinate": cell_coordinate,
            "evidence_kind": kind,
            "evidence_text": compact,
            "normalized_evidence": normalize_text(compact),
            "detected_domain": detected_domain,
            "occurrence_count": 0,
            "authority_implication": _authority_implication(kind),
            "nature": "observed_and_calculated",
        }
    accumulator[key]["occurrence_count"] = int(accumulator[key]["occurrence_count"]) + 1


def _profile_column(
    sheet: object,
    *,
    dataset_identity: str,
    raw_relative_path: str,
    header_row: int,
    column_number: int,
) -> dict[str, object]:
    header = sheet.cell(row=header_row, column=column_number).value
    values: list[object] = []
    formats: Counter[str] = Counter()
    semantic_types: Counter[str] = Counter()
    percent_style_cells = 0
    formula_cells = 0
    for row_number in range(header_row + 1, int(sheet.max_row) + 1):
        cell = sheet.cell(row=row_number, column=column_number)
        if cell.value is None:
            continue
        values.append(cell.value)
        semantic_types[canonical_value(cell.value)[0]] += 1
        number_format = str(cell.number_format or "General")
        formats[number_format] += 1
        percent_style_cells += int("%" in number_format)
        formula_cells += int(cell.data_type == "f")
    return {
        "dataset_identity": dataset_identity,
        "raw_relative_path": raw_relative_path,
        "sheet_name": str(sheet.title),
        "column_number": column_number,
        "header_raw": _text(header),
        "header_normalized": normalize_header(header),
        "unit_hint": _unit_hint(header),
        "nonempty_data_cells": len(values),
        "semantic_types": "|".join(sorted(semantic_types)),
        "semantic_type_counts_json": json.dumps(
            dict(sorted(semantic_types.items())),
            ensure_ascii=False,
            sort_keys=True,
        ),
        "excel_number_formats_json": json.dumps(
            dict(sorted(formats.items())),
            ensure_ascii=False,
            sort_keys=True,
        ),
        "percent_style_cells": percent_style_cells,
        "formula_cells": formula_cells,
        "unit_validation_status": "PENDING_SOURCE_METADATA_REVIEW",
        "nature": "observed_and_calculated_hint",
    }


def _manifest_lookup(manifest: pd.DataFrame) -> dict[str, object]:
    missing = MANIFEST_REQUIRED_COLUMNS.difference(manifest.columns)
    if missing:
        raise ValueError(
            "Colunas obrigatórias ausentes no manifesto da captura: "
            f"{sorted(missing)}"
        )
    if manifest["relative_path"].duplicated().any():
        raise ValueError("Manifesto da captura contém caminhos duplicados.")
    return {
        str(row.relative_path): row
        for row in manifest.itertuples(index=False)
    }


def audit_census_provenance(
    lineage: pd.DataFrame,
    source_manifest: pd.DataFrame,
    *,
    raw_snapshot_root: Path,
) -> CensusProvenanceAuditResult:
    """Extrai evidências locais sem atribuir autoridade institucional à fonte."""
    missing = LINEAGE_REQUIRED_COLUMNS.difference(lineage.columns)
    if missing:
        raise ValueError(
            "Colunas obrigatórias ausentes no registro de linhagem: "
            f"{sorted(missing)}"
        )
    matched = lineage.loc[lineage["lineage_match_status"].eq(MATCHED_STATUS)].copy()
    if matched.empty:
        raise ValueError("Nenhum par censitário um-para-um disponível para auditoria.")

    manifest_by_path = _manifest_lookup(source_manifest)
    root = raw_snapshot_root.expanduser().resolve()
    workbook_records: list[dict[str, object]] = []
    sheet_records: list[dict[str, object]] = []
    column_records: list[dict[str, object]] = []
    evidence_accumulator: dict[tuple[str, str, str, str], dict[str, object]] = {}

    for row in matched.sort_values("dataset_identity").itertuples(index=False):
        dataset_identity = str(row.dataset_identity)
        raw_relative_path = _single_path(row.raw_source_paths, row.raw_source_count)
        if raw_relative_path not in manifest_by_path:
            raise ValueError(
                "Fonte censitária ausente no manifesto da captura: "
                f"{raw_relative_path}"
            )
        manifest_row = manifest_by_path[raw_relative_path]
        path = _safe_snapshot_file(root, raw_relative_path)
        workbook = load_workbook(path, read_only=False, data_only=False)
        before_evidence = len(evidence_accumulator)
        document_properties: dict[str, str] = {}
        for field in DOCUMENT_PROPERTY_FIELDS:
            value = _iso(getattr(workbook.properties, field, ""))
            document_properties[field] = value
            if value:
                _add_evidence(
                    evidence_accumulator,
                    dataset_identity=dataset_identity,
                    raw_relative_path=raw_relative_path,
                    sheet_name="",
                    cell_coordinate="",
                    kind=f"DOCUMENT_PROPERTY_{normalize_header(field).upper()}",
                    evidence_text=value,
                )

        workbook_domains: set[str] = set()
        workbook_kinds: Counter[str] = Counter()
        for sheet in workbook.worksheets:
            header_row = _first_nonempty_row(sheet)
            nonempty_cells = 0
            formula_cells = 0
            comment_cells = 0
            hyperlink_cells = 0
            number_formats: set[str] = set()
            headers: list[str] = []
            if header_row is not None:
                headers = [
                    _text(sheet.cell(row=header_row, column=column).value)
                    for column in range(1, int(sheet.max_column) + 1)
                ]
                for column in range(1, int(sheet.max_column) + 1):
                    column_records.append(
                        _profile_column(
                            sheet,
                            dataset_identity=dataset_identity,
                            raw_relative_path=raw_relative_path,
                            header_row=header_row,
                            column_number=column,
                        )
                    )

            for sheet_row in sheet.iter_rows():
                for cell in sheet_row:
                    if cell.value is not None:
                        nonempty_cells += 1
                        formula_cells += int(cell.data_type == "f")
                        number_formats.add(str(cell.number_format or "General"))
                        cell_text = _cell_text(cell)
                        for url in URL_PATTERN.findall(cell_text):
                            domain = _domain(url)
                            workbook_domains.add(domain)
                            _add_evidence(
                                evidence_accumulator,
                                dataset_identity=dataset_identity,
                                raw_relative_path=raw_relative_path,
                                sheet_name=str(sheet.title),
                                cell_coordinate=str(cell.coordinate),
                                kind="CELL_URL",
                                evidence_text=url,
                                detected_domain=domain,
                            )
                        for kind in _evidence_kind(cell_text):
                            workbook_kinds[kind] += 1
                            _add_evidence(
                                evidence_accumulator,
                                dataset_identity=dataset_identity,
                                raw_relative_path=raw_relative_path,
                                sheet_name=str(sheet.title),
                                cell_coordinate=str(cell.coordinate),
                                kind=kind,
                                evidence_text=cell_text,
                            )
                    if cell.hyperlink is not None:
                        hyperlink_cells += 1
                        target = str(cell.hyperlink.target or cell.hyperlink.location or "")
                        domain = _domain(target) if target.lower().startswith(("http://", "https://")) else ""
                        if domain:
                            workbook_domains.add(domain)
                        _add_evidence(
                            evidence_accumulator,
                            dataset_identity=dataset_identity,
                            raw_relative_path=raw_relative_path,
                            sheet_name=str(sheet.title),
                            cell_coordinate=str(cell.coordinate),
                            kind="HYPERLINK_TARGET",
                            evidence_text=target,
                            detected_domain=domain,
                        )
                    if cell.comment is not None:
                        comment_cells += 1
                        comment_text = _text(cell.comment.text)
                        _add_evidence(
                            evidence_accumulator,
                            dataset_identity=dataset_identity,
                            raw_relative_path=raw_relative_path,
                            sheet_name=str(sheet.title),
                            cell_coordinate=str(cell.coordinate),
                            kind="CELL_COMMENT",
                            evidence_text=comment_text,
                        )
                        for kind in _evidence_kind(comment_text):
                            workbook_kinds[kind] += 1
                            _add_evidence(
                                evidence_accumulator,
                                dataset_identity=dataset_identity,
                                raw_relative_path=raw_relative_path,
                                sheet_name=str(sheet.title),
                                cell_coordinate=str(cell.coordinate),
                                kind=kind,
                                evidence_text=comment_text,
                            )

            sheet_records.append(
                {
                    "dataset_identity": dataset_identity,
                    "raw_relative_path": raw_relative_path,
                    "sheet_name": str(sheet.title),
                    "sheet_state": str(sheet.sheet_state),
                    "max_rows": int(sheet.max_row),
                    "max_columns": int(sheet.max_column),
                    "header_row_number": header_row or 0,
                    "headers_json": json.dumps(headers, ensure_ascii=False),
                    "nonempty_cells": nonempty_cells,
                    "data_rows_estimate": max(0, int(sheet.max_row) - (header_row or 0)),
                    "formula_cells": formula_cells,
                    "comment_cells": comment_cells,
                    "hyperlink_cells": hyperlink_cells,
                    "merged_range_count": len(sheet.merged_cells.ranges),
                    "hidden_row_count": sum(
                        bool(dimension.hidden)
                        for dimension in sheet.row_dimensions.values()
                    ),
                    "hidden_column_count": sum(
                        bool(dimension.hidden)
                        for dimension in sheet.column_dimensions.values()
                    ),
                    "excel_number_formats": "|".join(sorted(number_formats)),
                    "nature": "observed_and_calculated",
                }
            )

        evidence_added = len(evidence_accumulator) - before_evidence
        source_claim_count = int(workbook_kinds["SOURCE_LABEL"])
        institution_claim_count = int(workbook_kinds["INSTITUTION_LABEL"])
        if workbook_domains or source_claim_count or institution_claim_count:
            provenance_status = "EMBEDDED_PROVENANCE_EVIDENCE_DETECTED"
            authority_status = "PENDING_EXTERNAL_VERIFICATION"
        elif document_properties.get("creator") or document_properties.get(
            "lastModifiedBy"
        ):
            provenance_status = "DOCUMENT_METADATA_ONLY"
            authority_status = "NOT_ESTABLISHED"
        else:
            provenance_status = "NO_EMBEDDED_PROVENANCE_EVIDENCE"
            authority_status = "NOT_ESTABLISHED"

        workbook_records.append(
            {
                "dataset_identity": dataset_identity,
                "raw_relative_path": raw_relative_path,
                "file_size_bytes": path.stat().st_size,
                "manifest_expected_size_bytes": int(manifest_row.expected_size_bytes),
                "manifest_downloaded_size_bytes": int(
                    manifest_row.downloaded_size_bytes
                ),
                "manifest_verification_status": str(manifest_row.verification_status),
                "expected_sha256": str(manifest_row.expected_sha256 or ""),
                "local_sha256": str(manifest_row.local_sha256 or ""),
                "sheet_count": len(workbook.sheetnames),
                "sheet_names": "|".join(workbook.sheetnames),
                "creator": document_properties.get("creator", ""),
                "last_modified_by": document_properties.get("lastModifiedBy", ""),
                "created_at": _iso(workbook.properties.created),
                "modified_at": _iso(workbook.properties.modified),
                "document_title": document_properties.get("title", ""),
                "document_subject": document_properties.get("subject", ""),
                "document_description": document_properties.get("description", ""),
                "document_keywords": document_properties.get("keywords", ""),
                "defined_names_count": len(list(workbook.defined_names.values())),
                "external_links_count": len(getattr(workbook, "_external_links", [])),
                "evidence_records_detected": evidence_added,
                "detected_domains": "|".join(sorted(workbook_domains)),
                "source_label_occurrences": source_claim_count,
                "institution_label_occurrences": institution_claim_count,
                "period_label_occurrences": int(workbook_kinds["PERIOD_LABEL"]),
                "unit_label_occurrences": int(workbook_kinds["UNIT_LABEL"]),
                "geography_label_occurrences": int(
                    workbook_kinds["GEOGRAPHY_LABEL"]
                ),
                "provenance_status": provenance_status,
                "source_authority_status": authority_status,
                "conceptual_validation_status": "NOT_VALIDATED",
                "nature": "observed_and_calculated",
            }
        )
        workbook.close()

    workbooks = pd.DataFrame(workbook_records).sort_values(
        "dataset_identity"
    ).reset_index(drop=True)
    sheets = pd.DataFrame(sheet_records).sort_values(
        ["dataset_identity", "sheet_name"]
    ).reset_index(drop=True)
    columns = pd.DataFrame(column_records).sort_values(
        ["dataset_identity", "sheet_name", "column_number"]
    ).reset_index(drop=True)
    evidence = pd.DataFrame(list(evidence_accumulator.values()))
    if not evidence.empty:
        evidence = evidence.sort_values(
            ["dataset_identity", "evidence_kind", "evidence_text"]
        ).reset_index(drop=True)
    domains = {
        domain
        for value in workbooks["detected_domains"].fillna("")
        for domain in str(value).split("|")
        if domain
    }
    summary = pd.DataFrame(
        [
            ("workbooks_reviewed", len(workbooks), "observed"),
            (
                "manifest_verified_workbooks",
                int(workbooks["manifest_verification_status"].eq("VERIFIED").sum()),
                "observed",
            ),
            ("sheets_reviewed", len(sheets), "observed"),
            ("columns_profiled", len(columns), "observed"),
            ("evidence_records", len(evidence), "calculated"),
            (
                "workbooks_with_embedded_domains",
                int(workbooks["detected_domains"].ne("").sum()),
                "calculated",
            ),
            (
                "unique_detected_domains",
                len(domains),
                "calculated",
            ),
            (
                "workbooks_with_source_labels",
                int(workbooks["source_label_occurrences"].gt(0).sum()),
                "calculated",
            ),
            (
                "workbooks_with_document_creator",
                int(workbooks["creator"].ne("").sum()),
                "observed",
            ),
            (
                "workbooks_without_embedded_provenance",
                int(
                    workbooks["provenance_status"]
                    .eq("NO_EMBEDDED_PROVENANCE_EVIDENCE")
                    .sum()
                ),
                "calculated",
            ),
            ("source_authority_reviews_completed", 0, "observed"),
            ("conceptually_validated_datasets", 0, "observed"),
        ],
        columns=["indicator", "value", "nature"],
    )
    return CensusProvenanceAuditResult(
        workbooks=workbooks,
        sheets=sheets,
        columns=columns,
        evidence=evidence,
        summary=summary,
    )


def write_census_provenance_audit(
    result: CensusProvenanceAuditResult,
    output_dir: Path,
    *,
    replace: bool = False,
) -> Path:
    """Publica a auditoria local de forma atômica."""
    target = output_dir.expanduser().resolve()
    if target.exists():
        if not replace:
            raise FileExistsError(f"Destino da auditoria já existe: {target}")
        shutil.rmtree(target)
    partial = target.with_name(f".{target.name}.partial")
    if partial.exists():
        shutil.rmtree(partial)
    partial.mkdir(parents=True, exist_ok=False)
    outputs = {
        "demography_census_workbook_provenance.csv": result.workbooks,
        "demography_census_sheet_register.csv": result.sheets,
        "demography_census_column_metadata.csv": result.columns,
        "demography_census_provenance_evidence.csv": result.evidence,
        "demography_census_provenance_summary.csv": result.summary,
    }
    try:
        for file_name, frame in outputs.items():
            frame.to_csv(partial / file_name, index=False)
        target.parent.mkdir(parents=True, exist_ok=True)
        partial.rename(target)
    except Exception:
        shutil.rmtree(partial, ignore_errors=True)
        raise
    return target
