"""Perfil estrutural local dos arquivos capturados de ``raw/new_files``."""

from __future__ import annotations

import csv
import hashlib
import re
import unicodedata
from collections import Counter
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path, PurePosixPath

import pandas as pd
from openpyxl import load_workbook

SUPPORTED_EXCEL = {".xlsx", ".xlsm"}
SUPPORTED_DELIMITED = {".csv", ".tsv", ".txt"}
YEAR_PATTERN = re.compile(r"(?<!\d)(?:18|19|20|21)\d{2}(?!\d)")
INTEGER_PATTERN = re.compile(r"[-+]?\d+")
DECIMAL_PATTERN = re.compile(
    r"[-+]?(?:\d+[.,]\d+|\d{1,3}(?:\.\d{3})+,[0-9]+)"
)
DATE_TEXT_PATTERNS = (
    re.compile(r"\d{4}-\d{1,2}-\d{1,2}"),
    re.compile(r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}"),
)


@dataclass(frozen=True)
class CellObservation:
    value: object
    kind: str


@dataclass(frozen=True)
class ProfileResult:
    files: pd.DataFrame
    sheets: pd.DataFrame
    columns: pd.DataFrame
    schema_groups: pd.DataFrame


def normalize_label(value: object) -> str:
    """Normaliza um rótulo para comparação estrutural, sem alterar o original."""
    text = str(value or "").strip().lower()
    decomposed = unicodedata.normalize("NFKD", text)
    ascii_text = "".join(
        char for char in decomposed if not unicodedata.combining(char)
    )
    return re.sub(r"[^a-z0-9]+", "_", ascii_text).strip("_")


def _years_from_value(value: object) -> list[int]:
    if isinstance(value, (datetime, date)):
        return [value.year]
    return [
        int(match.group())
        for match in YEAR_PATTERN.finditer(str(value or ""))
    ]


def _classify_text(value: str) -> str:
    text = value.strip()
    if not text:
        return "blank"
    if any(pattern.fullmatch(text) for pattern in DATE_TEXT_PATTERNS):
        return "date_text"
    digits = text.lstrip("+-")
    has_preserved_leading_zero = len(digits) > 1 and digits.startswith("0")
    if INTEGER_PATTERN.fullmatch(text) and not has_preserved_leading_zero:
        return "integer_text"
    if DECIMAL_PATTERN.fullmatch(text):
        return "decimal_text"
    return "text"


def classify_value(
    value: object,
    *,
    data_type: str | None = None,
    is_date: bool = False,
) -> str:
    """Classifica o tipo aparente de uma célula observada."""
    if value is None or value == "":
        return "blank"
    if data_type == "f":
        return "formula"
    if data_type == "e":
        return "error"
    if is_date or isinstance(value, (datetime, date, time)):
        return "date"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int):
        return "integer"
    if isinstance(value, float):
        return "decimal"
    if isinstance(value, str):
        return _classify_text(value)
    return "other"


def _trim_row(row: list[CellObservation]) -> list[CellObservation]:
    last = 0
    for index, cell in enumerate(row, start=1):
        if cell.kind != "blank":
            last = index
    return row[:last]


def _header_candidate(
    rows: list[list[CellObservation]],
    max_scan_rows: int = 25,
) -> tuple[int | None, str]:
    candidates: list[tuple[tuple[float, ...], int, str]] = []
    for row_index, row in enumerate(rows[:max_scan_rows], start=1):
        trimmed = _trim_row(row)
        nonblank = [cell for cell in trimmed if cell.kind != "blank"]
        if not nonblank:
            continue
        text_like = sum(
            cell.kind in {"text", "date_text"} for cell in nonblank
        )
        normalized = [normalize_label(cell.value) for cell in nonblank]
        unique_ratio = len(set(normalized)) / len(normalized)
        text_ratio = text_like / len(nonblank)
        plausible = float(text_ratio >= 0.50 and len(nonblank) >= 2)
        score = (
            plausible,
            float(len(nonblank)),
            text_ratio,
            unique_ratio,
            float(-row_index),
        )
        if len(nonblank) >= 2 and text_ratio >= 0.75 and unique_ratio >= 0.75:
            confidence = "HIGH"
        elif len(nonblank) >= 2 and text_ratio >= 0.50:
            confidence = "MEDIUM"
        else:
            confidence = "LOW"
        candidates.append((score, row_index, confidence))

    if not candidates:
        return None, "NONE"
    _, row_index, confidence = max(candidates, key=lambda item: item[0])
    return row_index, confidence


def _header_values(
    rows: list[list[CellObservation]],
    header_row: int | None,
    observed_max_column: int,
) -> list[str]:
    if header_row is None:
        return [
            f"column_{index}"
            for index in range(1, observed_max_column + 1)
        ]
    row = rows[header_row - 1]
    values: list[str] = []
    for index in range(observed_max_column):
        value = row[index].value if index < len(row) else ""
        values.append(str(value or "").strip())
    return values


def _schema_signature(headers: list[str]) -> str:
    normalized = [
        normalize_label(value) or f"__blank_{index}"
        for index, value in enumerate(headers, start=1)
    ]
    return hashlib.sha256("|".join(normalized).encode("utf-8")).hexdigest()


def _inferred_type(counts: Counter[str]) -> str:
    usable = Counter(
        {
            kind: count
            for kind, count in counts.items()
            if kind not in {"blank", "formula", "error"} and count > 0
        }
    )
    if not usable:
        if counts.get("formula", 0):
            return "formula"
        return "empty"
    total = sum(usable.values())
    common = usable.most_common()
    if len(common) > 1 and common[1][1] / total >= 0.10:
        return "mixed"
    return common[0][0]


def _cell_at(
    row: list[CellObservation],
    column_index: int,
) -> CellObservation:
    if column_index <= len(row):
        return row[column_index - 1]
    return CellObservation(None, "blank")


def _profile_columns(
    *,
    relative_path: str,
    sheet_name: str,
    sheet_index: int,
    rows: list[list[CellObservation]],
    data_start: int,
    headers: list[str],
    normalized_headers: list[str],
    observed_max_column: int,
) -> tuple[list[dict[str, object]], list[int]]:
    records: list[dict[str, object]] = []
    all_years: list[int] = []

    for column_index in range(1, observed_max_column + 1):
        counts: Counter[str] = Counter()
        unique_values: set[str] = set()
        years: list[int] = []
        for row in rows[data_start:]:
            cell = _cell_at(row, column_index)
            counts[cell.kind] += 1
            if cell.kind != "blank":
                unique_values.add(f"{cell.kind}:{cell.value!s}")
                years.extend(_years_from_value(cell.value))
        all_years.extend(years)
        records.append(
            {
                "relative_path": relative_path,
                "sheet_name": sheet_name,
                "sheet_index": sheet_index,
                "column_index": column_index,
                "header_observed": headers[column_index - 1],
                "header_normalized": normalized_headers[column_index - 1],
                "nonempty_count": int(
                    sum(counts.values()) - counts.get("blank", 0)
                ),
                "blank_count": int(counts.get("blank", 0)),
                "unique_value_count_observed": len(unique_values),
                "inferred_type_estimate": _inferred_type(counts),
                "text_count": int(counts.get("text", 0)),
                "integer_count": int(counts.get("integer", 0)),
                "decimal_count": int(counts.get("decimal", 0)),
                "date_count": int(counts.get("date", 0)),
                "date_text_count": int(counts.get("date_text", 0)),
                "integer_text_count": int(counts.get("integer_text", 0)),
                "decimal_text_count": int(counts.get("decimal_text", 0)),
                "boolean_count": int(counts.get("boolean", 0)),
                "formula_count": int(counts.get("formula", 0)),
                "error_count": int(counts.get("error", 0)),
                "other_count": int(counts.get("other", 0)),
                "year_min_observed": min(years) if years else None,
                "year_max_observed": max(years) if years else None,
            }
        )
    return records, all_years


def _profile_table(
    *,
    relative_path: str,
    sheet_name: str,
    sheet_index: int,
    sheet_state: str,
    rows: list[list[CellObservation]],
    declared_max_row: int,
    declared_max_column: int,
    source_format: str,
    delimiter: str = "",
    encoding: str = "",
) -> tuple[dict[str, object], list[dict[str, object]]]:
    observed_nonempty_rows = 0
    observed_max_column = 0
    observed_nonempty_cells = 0
    first_nonempty_row: int | None = None
    last_nonempty_row: int | None = None

    for row_index, row in enumerate(rows, start=1):
        nonblank_positions = [
            index
            for index, cell in enumerate(row, start=1)
            if cell.kind != "blank"
        ]
        if not nonblank_positions:
            continue
        observed_nonempty_rows += 1
        observed_nonempty_cells += len(nonblank_positions)
        observed_max_column = max(
            observed_max_column,
            max(nonblank_positions),
        )
        first_nonempty_row = first_nonempty_row or row_index
        last_nonempty_row = row_index

    header_row, header_confidence = _header_candidate(rows)
    headers = _header_values(rows, header_row, observed_max_column)
    normalized_headers = [normalize_label(value) for value in headers]
    nonblank_headers = [value for value in normalized_headers if value]
    duplicate_header_labels = sum(
        count - 1
        for count in Counter(nonblank_headers).values()
        if count > 1
    )
    blank_header_cells = sum(not value for value in normalized_headers)
    signature = _schema_signature(headers) if observed_max_column else ""
    data_start = header_row if header_row is not None else 0

    column_records, all_years = _profile_columns(
        relative_path=relative_path,
        sheet_name=sheet_name,
        sheet_index=sheet_index,
        rows=rows,
        data_start=data_start,
        headers=headers,
        normalized_headers=normalized_headers,
        observed_max_column=observed_max_column,
    )

    sheet_record = {
        "relative_path": relative_path,
        "source_format": source_format,
        "sheet_name": sheet_name,
        "sheet_index": sheet_index,
        "sheet_state": sheet_state,
        "delimiter": delimiter,
        "encoding": encoding,
        "declared_max_row": declared_max_row,
        "declared_max_column": declared_max_column,
        "observed_nonempty_rows": observed_nonempty_rows,
        "observed_nonempty_cells": observed_nonempty_cells,
        "observed_max_column": observed_max_column,
        "first_nonempty_row": first_nonempty_row,
        "last_nonempty_row": last_nonempty_row,
        "header_row_candidate_estimate": header_row,
        "header_confidence_estimate": header_confidence,
        "header_detection_method": "HEURISTIC_FIRST_25_ROWS",
        "header_nonblank_cells": len(nonblank_headers),
        "header_blank_cells": blank_header_cells,
        "header_duplicate_labels": duplicate_header_labels,
        "schema_signature_sha256": signature,
        "year_min_observed": min(all_years) if all_years else None,
        "year_max_observed": max(all_years) if all_years else None,
    }
    return sheet_record, column_records


def _excel_tables(
    path: Path,
    relative_path: str,
) -> Iterable[tuple[dict[str, object], list[dict[str, object]]]]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        for sheet_index, worksheet in enumerate(
            workbook.worksheets,
            start=1,
        ):
            rows: list[list[CellObservation]] = []
            for source_row in worksheet.iter_rows():
                row = [
                    CellObservation(
                        cell.value,
                        classify_value(
                            cell.value,
                            data_type=getattr(cell, "data_type", None),
                            is_date=bool(getattr(cell, "is_date", False)),
                        ),
                    )
                    for cell in source_row
                ]
                rows.append(_trim_row(row))
            yield _profile_table(
                relative_path=relative_path,
                sheet_name=worksheet.title,
                sheet_index=sheet_index,
                sheet_state=worksheet.sheet_state,
                rows=rows,
                declared_max_row=int(worksheet.max_row or 0),
                declared_max_column=int(worksheet.max_column or 0),
                source_format=path.suffix.lower().lstrip("."),
            )
    finally:
        workbook.close()


def _decode_delimited(path: Path) -> tuple[str, str]:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return raw.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    raise ValueError(f"Não foi possível decodificar o arquivo: {path}")


def _delimited_table(
    path: Path,
    relative_path: str,
) -> tuple[dict[str, object], list[dict[str, object]]]:
    text, encoding = _decode_delimited(path)
    sample = text[:65536]
    default_delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = default_delimiter

    parsed = list(csv.reader(text.splitlines(), delimiter=delimiter))
    rows = [
        _trim_row(
            [
                CellObservation(value, classify_value(value))
                for value in source_row
            ]
        )
        for source_row in parsed
    ]
    declared_max_column = max((len(row) for row in rows), default=0)
    return _profile_table(
        relative_path=relative_path,
        sheet_name="(delimited_file)",
        sheet_index=1,
        sheet_state="visible",
        rows=rows,
        declared_max_row=len(rows),
        declared_max_column=declared_max_column,
        source_format=path.suffix.lower().lstrip("."),
        delimiter="\\t" if delimiter == "\t" else delimiter,
        encoding=encoding,
    )


def _empty_schema_groups() -> pd.DataFrame:
    return pd.DataFrame(
        columns=[
            "schema_signature_sha256",
            "group_size",
            "relative_path",
            "sheet_name",
            "header_confidence_estimate",
        ]
    )


def _schema_groups(sheets: pd.DataFrame) -> pd.DataFrame:
    if sheets.empty:
        return _empty_schema_groups()
    eligible = sheets.loc[
        sheets["schema_signature_sha256"].fillna("").astype(str).ne("")
        & sheets["header_confidence_estimate"].isin(["HIGH", "MEDIUM"])
    ].copy()
    repeated = eligible.loc[
        eligible.duplicated("schema_signature_sha256", keep=False),
        [
            "schema_signature_sha256",
            "relative_path",
            "sheet_name",
            "header_confidence_estimate",
        ],
    ].copy()
    if repeated.empty:
        return _empty_schema_groups()
    repeated.insert(
        1,
        "group_size",
        repeated.groupby("schema_signature_sha256")["relative_path"]
        .transform("size")
        .astype(int),
    )
    return repeated.sort_values(
        [
            "group_size",
            "schema_signature_sha256",
            "relative_path",
            "sheet_name",
        ],
        ascending=[False, True, True, True],
    ).reset_index(drop=True)


def profile_snapshot(snapshot_path: Path) -> ProfileResult:
    """Perfila os arquivos sob ``raw/new_files`` sem modificar a captura."""
    root = snapshot_path.expanduser().resolve()
    inbox_root = root / "raw" / "new_files"
    if not inbox_root.is_dir():
        raise FileNotFoundError(
            f"Caixa capturada não encontrada: {inbox_root}"
        )

    file_records: list[dict[str, object]] = []
    sheet_records: list[dict[str, object]] = []
    column_records: list[dict[str, object]] = []
    paths = sorted(
        candidate
        for candidate in inbox_root.rglob("*")
        if candidate.is_file()
    )

    for path in paths:
        relative_path = PurePosixPath(
            *path.relative_to(root).parts
        ).as_posix()
        extension = path.suffix.lower()
        file_record: dict[str, object] = {
            "relative_path": relative_path,
            "extension": extension.lstrip("."),
            "size_bytes": path.stat().st_size,
            "profile_status": "PENDING",
            "sheet_count": 0,
            "error_type": "",
            "error_message": "",
        }
        try:
            if extension in SUPPORTED_EXCEL:
                tables = list(_excel_tables(path, relative_path))
            elif extension in SUPPORTED_DELIMITED:
                tables = [_delimited_table(path, relative_path)]
            else:
                file_record["profile_status"] = "UNSUPPORTED_FORMAT"
                file_records.append(file_record)
                continue

            for sheet_record, columns in tables:
                sheet_records.append(sheet_record)
                column_records.extend(columns)
            file_record["profile_status"] = "PROFILED"
            file_record["sheet_count"] = len(tables)
        except Exception as exc:  # noqa: BLE001
            file_record["profile_status"] = "ERROR"
            file_record["error_type"] = type(exc).__name__
            file_record["error_message"] = str(exc)[:500]
        file_records.append(file_record)

    files = pd.DataFrame(file_records)
    sheets = pd.DataFrame(sheet_records)
    columns = pd.DataFrame(column_records)
    return ProfileResult(
        files=files,
        sheets=sheets,
        columns=columns,
        schema_groups=_schema_groups(sheets),
    )
