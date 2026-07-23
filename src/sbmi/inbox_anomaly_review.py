"""Revisão de anomalias da captura local de ``raw/new_files``."""

from __future__ import annotations

import csv
import hashlib
import json
import re
from collections import Counter, defaultdict
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import date, datetime
from itertools import combinations
from pathlib import Path, PurePosixPath

import pandas as pd
from openpyxl import load_workbook

from sbmi.inbox_content_audit import canonical_value, normalized_value
from sbmi.inbox_profile import normalize_label
from sbmi.inbox_structure_triage import source_from_path

TABLE_KEY = ["relative_path", "sheet_name", "sheet_index"]
DATE_HEADERS = (
    "mes_ano",
    "data",
    "competencia",
    "ano_mes",
    "mes_competencia",
    "data_pagamento",
    "data_lancamento",
)
COPY_SUFFIX_PATTERN = re.compile(r"\(\d+\)$")
ISO_DATE_PATTERN = re.compile(r"^(\d{4})-(\d{2})-(\d{2})$")
ISO_DATETIME_PATTERN = re.compile(r"^(\d{4})-(\d{2})-(\d{2})[ T](\d{2}):(\d{2}):(\d{2})$")
SLASH_DATE_PATTERN = re.compile(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{4})$")
MONTH_YEAR_PATTERN = re.compile(r"^(\d{1,2})[/-](\d{4})$")
YEAR_MONTH_PATTERN = re.compile(r"^(\d{4})[/-](\d{1,2})$")


@dataclass(frozen=True)
class ReviewResult:
    content_duplicate_pairs: pd.DataFrame
    duplicate_row_groups: pd.DataFrame
    temporal_table_summary: pd.DataFrame
    temporal_anomalies: pd.DataFrame
    review_summary: pd.DataFrame


@dataclass(frozen=True)
class RowObservation:
    source_row_number: int
    values: tuple[object, ...]


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _row_hash(values: Iterable[object], *, normalized: bool) -> str:
    converter = normalized_value if normalized else canonical_value
    payload = "\x1f".join(converter(value) for value in values)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _content_fingerprint(headers: Iterable[str], rows: Iterable[tuple[object, ...]]) -> str:
    header_payload = "\x1f".join(headers)
    row_hashes = sorted(_row_hash(row, normalized=True) for row in rows)
    payload = header_payload + "\n" + "\n".join(row_hashes)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _copy_suffix_rank(path_value: str) -> int:
    stem = PurePosixPath(path_value).stem.strip()
    return 1 if COPY_SUFFIX_PATTERN.search(stem) else 0


def _choose_primary(left_path: str, right_path: str) -> tuple[str, str, str]:
    left_rank = _copy_suffix_rank(left_path)
    right_rank = _copy_suffix_rank(right_path)
    if left_rank != right_rank:
        primary = left_path if left_rank < right_rank else right_path
        duplicate = right_path if primary == left_path else left_path
        return primary, duplicate, "COPY_SUFFIX_HEURISTIC"
    return "", "", "PENDING_MANUAL_REVIEW"


def build_content_duplicate_pairs(snapshot_path: Path, tables: Iterable[object]) -> pd.DataFrame:
    """Localiza conteúdos normalizados idênticos e compara o binário local."""
    root = snapshot_path.expanduser().resolve()
    grouped: dict[str, list[object]] = defaultdict(list)
    for table in tables:
        fingerprint = _content_fingerprint(table.normalized_headers, table.rows)
        grouped[fingerprint].append(table)

    records: list[dict[str, object]] = []
    for fingerprint, group in grouped.items():
        if len(group) < 2:
            continue
        for left, right in combinations(group, 2):
            left_file = root / PurePosixPath(left.relative_path)
            right_file = root / PurePosixPath(right.relative_path)
            left_binary = _sha256_file(left_file)
            right_binary = _sha256_file(right_file)
            primary, duplicate, basis = _choose_primary(
                left.relative_path,
                right.relative_path,
            )
            binary_same = left_binary == right_binary
            records.append(
                {
                    "content_fingerprint_sha256": fingerprint,
                    "left_path": left.relative_path,
                    "right_path": right.relative_path,
                    "left_source": left.source_declared,
                    "right_source": right.source_declared,
                    "rows_left": len(left.rows),
                    "rows_right": len(right.rows),
                    "binary_sha256_left": left_binary,
                    "binary_sha256_right": right_binary,
                    "binary_same": binary_same,
                    "duplicate_class": "EXACT_DUPLICATE"
                    if binary_same
                    else "CONTENT_DUPLICATE",
                    "suggested_primary_path": primary,
                    "suggested_duplicate_path": duplicate,
                    "suggestion_basis": basis,
                    "review_status": "PENDING_MANUAL_DISPOSITION",
                }
            )
    columns = [
        "content_fingerprint_sha256",
        "left_path",
        "right_path",
        "left_source",
        "right_source",
        "rows_left",
        "rows_right",
        "binary_sha256_left",
        "binary_sha256_right",
        "binary_same",
        "duplicate_class",
        "suggested_primary_path",
        "suggested_duplicate_path",
        "suggestion_basis",
        "review_status",
    ]
    if not records:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame(records).sort_values(
        ["duplicate_class", "left_path", "right_path"]
    ).reset_index(drop=True)


def _decode_delimited(path: Path) -> tuple[str, str]:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return raw.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError("utf-8", raw, 0, 1, "Não foi possível decodificar o arquivo")


def _table_rows(
    snapshot_path: Path,
    profile_row: object,
) -> tuple[list[str], list[RowObservation]]:
    path = snapshot_path / PurePosixPath(str(profile_row.relative_path))
    header_row = int(profile_row.header_row_candidate_estimate)
    width = int(profile_row.observed_max_column)
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            worksheet = workbook[str(profile_row.sheet_name)]
            headers: list[str] = []
            observations: list[RowObservation] = []
            for source_row_number, cells in enumerate(
                worksheet.iter_rows(max_col=width),
                start=1,
            ):
                values = tuple(cell.value for cell in cells)
                if source_row_number == header_row:
                    headers = [canonical_value(value) for value in values]
                elif source_row_number > header_row and any(
                    canonical_value(value) for value in values
                ):
                    observations.append(RowObservation(source_row_number, values))
            return headers, observations
        finally:
            workbook.close()

    text, _ = _decode_delimited(path)
    sample = text[:65536]
    default = "\t" if path.suffix.lower() == ".tsv" else ","
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        delimiter = default
    parsed = list(csv.reader(text.splitlines(), delimiter=delimiter))
    headers = [canonical_value(value) for value in parsed[header_row - 1][:width]]
    observations = []
    for source_row_number, row in enumerate(parsed[header_row:], start=header_row + 1):
        values = tuple(row[:width] + [None] * max(0, width - len(row)))
        if any(canonical_value(value) for value in values):
            observations.append(RowObservation(source_row_number, values))
    return headers, observations


def build_duplicate_row_groups(
    snapshot_path: Path,
    sheet_profile: pd.DataFrame,
) -> pd.DataFrame:
    """Detalha grupos de linhas repetidas, preservando os números de origem."""
    records: list[dict[str, object]] = []
    for profile_row in sheet_profile.itertuples(index=False):
        headers, observations = _table_rows(snapshot_path, profile_row)
        groups: dict[str, list[RowObservation]] = defaultdict(list)
        for observation in observations:
            groups[_row_hash(observation.values, normalized=True)].append(observation)
        for normalized_hash, group in groups.items():
            if len(group) < 2:
                continue
            strict_hashes = {_row_hash(item.values, normalized=False) for item in group}
            row_map = {
                normalize_label(header) or f"column_{index}": canonical_value(value)
                for index, (header, value) in enumerate(
                    zip(headers, group[0].values, strict=False),
                    start=1,
                )
            }
            records.append(
                {
                    "relative_path": str(profile_row.relative_path),
                    "source_declared": source_from_path(profile_row.relative_path),
                    "sheet_name": str(profile_row.sheet_name),
                    "sheet_index": int(profile_row.sheet_index),
                    "normalized_row_hash": normalized_hash,
                    "occurrence_count": len(group),
                    "duplicate_excess": len(group) - 1,
                    "source_row_numbers": "|".join(
                        str(item.source_row_number) for item in group
                    ),
                    "strict_variants": len(strict_hashes),
                    "duplicate_class": "STRICT_EXACT_ROW"
                    if len(strict_hashes) == 1
                    else "NORMALIZED_EQUIVALENT_ROW",
                    "row_values_json": json.dumps(row_map, ensure_ascii=False),
                    "review_status": "PENDING_SOURCE_VALIDATION",
                }
            )
    columns = [
        "relative_path",
        "source_declared",
        "sheet_name",
        "sheet_index",
        "normalized_row_hash",
        "occurrence_count",
        "duplicate_excess",
        "source_row_numbers",
        "strict_variants",
        "duplicate_class",
        "row_values_json",
        "review_status",
    ]
    if not records:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame(records).sort_values(
        ["duplicate_excess", "relative_path", "source_row_numbers"],
        ascending=[False, True, True],
    ).reset_index(drop=True)


def _safe_date(year: int, month: int, day: int) -> date | None:
    try:
        return date(year, month, day)
    except ValueError:
        return None


def parse_date_observation(value: object) -> dict[str, object]:
    """Interpreta uma data por formatos explícitos e registra ambiguidades."""
    if isinstance(value, datetime):
        return {
            "raw_type": "datetime",
            "raw_text": value.isoformat(sep=" "),
            "parsed_date": value.date(),
            "alternative_date": None,
            "parse_method": "PYTHON_DATETIME",
            "ambiguous": False,
        }
    if isinstance(value, date):
        return {
            "raw_type": "date",
            "raw_text": value.isoformat(),
            "parsed_date": value,
            "alternative_date": None,
            "parse_method": "PYTHON_DATE",
            "ambiguous": False,
        }

    text = canonical_value(value)
    base = {
        "raw_type": type(value).__name__,
        "raw_text": text,
        "parsed_date": None,
        "alternative_date": None,
        "parse_method": "UNPARSED",
        "ambiguous": False,
    }
    if not text:
        return base

    match = ISO_DATETIME_PATTERN.fullmatch(text)
    if match:
        parsed = _safe_date(int(match[1]), int(match[2]), int(match[3]))
        return {**base, "parsed_date": parsed, "parse_method": "ISO_DATETIME"}

    match = ISO_DATE_PATTERN.fullmatch(text)
    if match:
        parsed = _safe_date(int(match[1]), int(match[2]), int(match[3]))
        return {**base, "parsed_date": parsed, "parse_method": "ISO_DATE"}

    match = MONTH_YEAR_PATTERN.fullmatch(text)
    if match:
        parsed = _safe_date(int(match[2]), int(match[1]), 1)
        return {**base, "parsed_date": parsed, "parse_method": "MONTH_YEAR"}

    match = YEAR_MONTH_PATTERN.fullmatch(text)
    if match:
        parsed = _safe_date(int(match[1]), int(match[2]), 1)
        return {**base, "parsed_date": parsed, "parse_method": "YEAR_MONTH"}

    match = SLASH_DATE_PATTERN.fullmatch(text)
    if match:
        first, second, year = (int(match[1]), int(match[2]), int(match[3]))
        day_first = _safe_date(year, second, first)
        month_first = _safe_date(year, first, second)
        ambiguous = (
            day_first is not None
            and month_first is not None
            and day_first != month_first
        )
        return {
            **base,
            "parsed_date": day_first,
            "alternative_date": month_first if ambiguous else None,
            "parse_method": "DMY_ASSUMED",
            "ambiguous": ambiguous,
        }
    return base


def build_temporal_review(
    snapshot_path: Path,
    sheet_profile: pd.DataFrame,
    *,
    snapshot_date: date,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Resume colunas temporais e registra valores futuros, ambíguos ou inválidos."""
    summary_records: list[dict[str, object]] = []
    anomaly_records: list[dict[str, object]] = []
    for profile_row in sheet_profile.itertuples(index=False):
        headers, observations = _table_rows(snapshot_path, profile_row)
        normalized_headers = [normalize_label(value) for value in headers]
        date_index = next(
            (normalized_headers.index(name) for name in DATE_HEADERS if name in normalized_headers),
            None,
        )
        if date_index is None:
            continue

        parsed_dates: list[date] = []
        future_count = 0
        ambiguous_count = 0
        reversal_count = 0
        failure_count = 0
        raw_type_counter: Counter[str] = Counter()
        for observation in observations:
            value = observation.values[date_index]
            parsed = parse_date_observation(value)
            raw_type_counter[str(parsed["raw_type"])] += 1
            parsed_date = parsed["parsed_date"]
            alternative_date = parsed["alternative_date"]
            ambiguous = bool(parsed["ambiguous"])
            if parsed_date is None:
                failure_count += 1
                anomaly_class = "PARSE_FAILURE"
            else:
                parsed_dates.append(parsed_date)
                future = parsed_date > snapshot_date
                alternative_not_future = (
                    isinstance(alternative_date, date)
                    and alternative_date <= snapshot_date
                )
                if ambiguous and future and alternative_not_future:
                    anomaly_class = "AMBIGUOUS_DATE_POSSIBLE_REVERSAL"
                    reversal_count += 1
                elif future:
                    anomaly_class = "FUTURE_DATE"
                elif ambiguous:
                    anomaly_class = "AMBIGUOUS_DATE"
                else:
                    anomaly_class = ""
                future_count += int(future)
                ambiguous_count += int(ambiguous)
            if anomaly_class:
                anomaly_records.append(
                    {
                        "relative_path": str(profile_row.relative_path),
                        "source_declared": source_from_path(profile_row.relative_path),
                        "sheet_name": str(profile_row.sheet_name),
                        "source_row_number": observation.source_row_number,
                        "date_header": normalized_headers[date_index],
                        "raw_type": parsed["raw_type"],
                        "raw_text": parsed["raw_text"],
                        "parsed_date": parsed_date.isoformat()
                        if isinstance(parsed_date, date)
                        else "",
                        "alternative_date": alternative_date.isoformat()
                        if isinstance(alternative_date, date)
                        else "",
                        "parse_method": parsed["parse_method"],
                        "anomaly_class": anomaly_class,
                        "snapshot_date": snapshot_date.isoformat(),
                        "review_status": "PENDING_TEMPORAL_VALIDATION",
                    }
                )

        summary_records.append(
            {
                "relative_path": str(profile_row.relative_path),
                "source_declared": source_from_path(profile_row.relative_path),
                "sheet_name": str(profile_row.sheet_name),
                "date_header": normalized_headers[date_index],
                "values_observed": len(observations),
                "values_parsed": len(parsed_dates),
                "parse_failures": failure_count,
                "ambiguous_values": ambiguous_count,
                "future_values": future_count,
                "possible_reversal_values": reversal_count,
                "period_min_observed": min(parsed_dates).isoformat()
                if parsed_dates
                else "",
                "period_max_observed": max(parsed_dates).isoformat()
                if parsed_dates
                else "",
                "raw_types": "|".join(
                    f"{name}:{count}" for name, count in sorted(raw_type_counter.items())
                ),
            }
        )

    summary = pd.DataFrame(summary_records)
    anomalies = pd.DataFrame(anomaly_records)
    return summary, anomalies


def build_review_summary(
    content_duplicates: pd.DataFrame,
    duplicate_rows: pd.DataFrame,
    temporal_summary: pd.DataFrame,
    temporal_anomalies: pd.DataFrame,
) -> pd.DataFrame:
    """Produz indicadores agregados com natureza explicitada."""
    anomaly_classes = (
        temporal_anomalies["anomaly_class"].value_counts()
        if not temporal_anomalies.empty
        else {}
    )
    indicators = [
        ("content_duplicate_pairs", len(content_duplicates), "calculated"),
        (
            "content_duplicate_binary_different_pairs",
            int(content_duplicates["duplicate_class"].eq("CONTENT_DUPLICATE").sum())
            if not content_duplicates.empty
            else 0,
            "calculated",
        ),
        ("duplicate_row_groups", len(duplicate_rows), "calculated"),
        (
            "duplicate_row_excess",
            int(duplicate_rows["duplicate_excess"].sum())
            if not duplicate_rows.empty
            else 0,
            "calculated",
        ),
        (
            "tables_with_duplicate_rows",
            int(duplicate_rows["relative_path"].nunique())
            if not duplicate_rows.empty
            else 0,
            "calculated",
        ),
        ("temporal_tables", len(temporal_summary), "observed"),
        (
            "future_date_values",
            int(temporal_summary["future_values"].sum())
            if not temporal_summary.empty
            else 0,
            "calculated",
        ),
        (
            "ambiguous_date_values",
            int(temporal_summary["ambiguous_values"].sum())
            if not temporal_summary.empty
            else 0,
            "calculated",
        ),
        (
            "possible_date_reversal_values",
            int(anomaly_classes.get("AMBIGUOUS_DATE_POSSIBLE_REVERSAL", 0)),
            "calculated",
        ),
        (
            "date_parse_failures",
            int(temporal_summary["parse_failures"].sum())
            if not temporal_summary.empty
            else 0,
            "calculated",
        ),
    ]
    return pd.DataFrame(indicators, columns=["indicator", "value", "nature"])


def review_snapshot_anomalies(
    snapshot_path: Path,
    sheet_profile: pd.DataFrame,
    tables: Iterable[object],
    *,
    snapshot_date: date,
) -> ReviewResult:
    """Executa revisão de anomalias sem alterar a captura."""
    content_duplicates = build_content_duplicate_pairs(snapshot_path, tables)
    duplicate_rows = build_duplicate_row_groups(snapshot_path, sheet_profile)
    temporal_summary, temporal_anomalies = build_temporal_review(
        snapshot_path,
        sheet_profile,
        snapshot_date=snapshot_date,
    )
    review_summary = build_review_summary(
        content_duplicates,
        duplicate_rows,
        temporal_summary,
        temporal_anomalies,
    )
    return ReviewResult(
        content_duplicate_pairs=content_duplicates,
        duplicate_row_groups=duplicate_rows,
        temporal_table_summary=temporal_summary,
        temporal_anomalies=temporal_anomalies,
        review_summary=review_summary,
    )
